import mysql
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QMainWindow, QStyledItemDelegate, QAction
import sys
import os
import common as cm
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from doctor_manage import Ui_doctor_manage
from patient_manage import Ui_patient_manage
from schedule_manage import Ui_schedule_manage
from edit_hospital import Ui_edit_form_hospital


# data = cm.select_data('hospital')


class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return


class UI_hospital_manage(QMainWindow):
    def __init__(self, parent=None):
        super(UI_hospital_manage, self).__init__(parent)
        self.database_config_obj = None

    def setupUi(self):
        self.setObjectName("Hospital_Manage")
        self.resize(983, 762)
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(390, 0, 121, 51))
        self.setWindowIcon(QtGui.QIcon("./icon/logo.png"))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.table_hospitals = QtWidgets.QTableWidget(self.centralwidget)
        delegate = ReadOnlyDelegate(self)
        self.table_hospitals.setItemDelegateForColumn(0, delegate)
        self.table_hospitals.setItemDelegateForColumn(1, delegate)
        self.table_hospitals.setItemDelegateForColumn(2, delegate)
        self.table_hospitals.setItemDelegateForColumn(3, delegate)
        self.table_hospitals.setItemDelegateForColumn(4, delegate)
        self.table_hospitals.setGeometry(QtCore.QRect(30, 50, 831, 280))
        self.table_hospitals.setMinimumSize(QtCore.QSize(0, 300))
        self.table_hospitals.setObjectName("table_hospitals")
        self.table_hospitals.setColumnCount(7)
        self.table_hospitals.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_hospitals.setHorizontalHeaderItem(6, item)
        self.btn_import = QtWidgets.QPushButton(self.centralwidget)
        self.btn_import.setToolTip('User should be use template')
        self.btn_import.clicked.connect(self.import_excel)
        self.btn_import.setGeometry(QtCore.QRect(870, 70, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_import.setFont(font)
        self.btn_import.setObjectName("btn_import")
        self.btn_export = QtWidgets.QPushButton(self.centralwidget)
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_export.setGeometry(QtCore.QRect(870, 130, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_export.setFont(font)
        self.btn_export.setObjectName("btn_export")
        self.tab_search_add = QtWidgets.QTabWidget(self.centralwidget)
        self.tab_search_add.setGeometry(QtCore.QRect(30, 370, 941, 311))
        self.tab_search_add.setObjectName("tab_search_add")
        self.tab_search = QtWidgets.QWidget()
        self.tab_search.setObjectName("tab_search")
        self.groupBox = QtWidgets.QGroupBox(self.tab_search)
        self.groupBox.setGeometry(QtCore.QRect(20, 10, 811, 141))
        self.groupBox.setObjectName("groupBox")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(40, 40, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.txt_search = QtWidgets.QLineEdit(self.groupBox)
        self.txt_search.setGeometry(QtCore.QRect(130, 30, 661, 31))
        self.txt_search.setObjectName("txt_search")
        self.btn_search = QtWidgets.QPushButton(self.groupBox)
        self.btn_search.clicked.connect(self.search_hospital)
        self.btn_search.setGeometry(QtCore.QRect(330, 90, 151, 31))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.btn_search.setFont(font)
        self.btn_search.setObjectName("btn_search")
        self.tab_search_add.addTab(self.tab_search, "")
        self.tab_add_new = QtWidgets.QWidget()
        self.tab_add_new.setObjectName("tab_add_new")
        self.label_3 = QtWidgets.QLabel(self.tab_add_new)
        self.label_3.setGeometry(QtCore.QRect(40, 40, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.tab_add_new)
        self.label_4.setGeometry(QtCore.QRect(40, 90, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.tab_add_new)
        self.label_5.setGeometry(QtCore.QRect(40, 140, 71, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.tab_add_new)
        self.label_6.setGeometry(QtCore.QRect(40, 190, 101, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.txt_name = QtWidgets.QLineEdit(self.tab_add_new)
        self.txt_name.setGeometry(QtCore.QRect(180, 31, 621, 31))
        self.txt_name.setObjectName("txt_name")
        self.txt_phone = QtWidgets.QLineEdit(self.tab_add_new)
        self.txt_phone.setGeometry(QtCore.QRect(180, 80, 621, 31))
        self.txt_phone.setObjectName("txt_phone")
        self.txt_address = QtWidgets.QLineEdit(self.tab_add_new)
        self.txt_address.setGeometry(QtCore.QRect(180, 130, 621, 31))
        self.txt_address.setObjectName("txt_address")
        self.txt_description = QtWidgets.QLineEdit(self.tab_add_new)
        self.txt_description.setGeometry(QtCore.QRect(180, 180, 621, 31))
        self.txt_description.setObjectName("txt_description")
        self.btn_add = QtWidgets.QPushButton(self.tab_add_new)
        self.btn_add.clicked.connect(self.add_new)
        self.btn_add.setGeometry(QtCore.QRect(370, 230, 89, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btn_add.setFont(font)
        self.btn_add.setObjectName("btn_add")
        self.tab_search_add.addTab(self.tab_add_new, "")
        self.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 983, 25))
        self.menubar.setObjectName("menubar")
        self.menuSetting = QtWidgets.QMenu(self.menubar)
        self.menuSetting.setObjectName("menuSetting")
        self.menuPrograms = QtWidgets.QMenu(self.menubar)
        self.menuPrograms.setObjectName("menuPrograms")
        self.setMenuBar(self.menubar)
        self.action_config = QtWidgets.QAction(self)
        self.action_config = QAction(QtGui.QIcon("./icon/config.png"), "&Your button", self)
        self.action_config.setObjectName("action_config")
        # self.action_config.triggered.connect(self.open_notepad_config)
        self.actionPatient = QtWidgets.QAction(self)
        self.actionPatient.setObjectName("actionPatient")
        self.actionSchedule = QtWidgets.QAction(self)
        self.actionSchedule.setObjectName("actionSchedule")
        self.action_doctor = QtWidgets.QAction(self)
        self.action_doctor = QAction(QtGui.QIcon("./icon/setting_3.png"), "Your &button2", self)
        self.action_doctor.setObjectName("action_doctor")
        self.action_doctor.triggered.connect(self.open_doctor_manage)
        self.action_patient = QtWidgets.QAction(self)
        self.action_patient = QAction(QtGui.QIcon("./icon/setting_3.png"), "Your &button2", self)
        self.action_patient.triggered.connect(self.open_patient_manage)
        self.action_patient.setObjectName("action_patient")
        self.action_schedule = QtWidgets.QAction(self)
        self.action_schedule = QAction(QtGui.QIcon("./icon/setting_3.png"), "Your &button2", self)
        self.action_schedule.triggered.connect(self.open_schedule)
        self.action_schedule.setObjectName("action_schedule")
        self.menuSetting.addAction(self.action_config)
        self.menuPrograms.addAction(self.action_doctor)
        self.menuPrograms.addAction(self.action_patient)
        self.menuPrograms.addAction(self.action_schedule)
        self.menubar.addAction(self.menuSetting.menuAction())
        self.menubar.addAction(self.menuPrograms.menuAction())
        self.retranslateUi()
        self.tab_search_add.setCurrentIndex(0)
        try:
            self.database_config_obj = cm.load_config(cm.DATABASE_CONFIG_PATH)
            self.data_list = self.db_get_data_list(self.database_config_obj)
        except Exception as e:
            print(e)
            QMessageBox.information(self, 'Config Message', f'{e}Your config database file not suitable, try again',
                                    QMessageBox.Close)
            self.data_list = []
        self.action_config.triggered.connect(self.open_notepad_config)
        if self.data_list:
            self.init_data(self.data_list)
        else:
            return

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Hospital_Manage", "Hospital Manage"))
        self.label.setText(_translate("Hospital_Manage", "Hospitals"))
        item = self.table_hospitals.horizontalHeaderItem(0)
        item.setText(_translate("Hospital_Manage", "ID"))
        item = self.table_hospitals.horizontalHeaderItem(1)
        item.setText(_translate("Hospital_Manage", "Name"))
        item = self.table_hospitals.horizontalHeaderItem(2)
        item.setText(_translate("Hospital_Manage", "Phone"))
        item = self.table_hospitals.horizontalHeaderItem(3)
        item.setText(_translate("Hospital_Manage", "Address"))
        item = self.table_hospitals.horizontalHeaderItem(4)
        item.setText(_translate("Hospital_Manage", "Description"))
        item = self.table_hospitals.horizontalHeaderItem(5)
        item.setText(_translate("Hospital_Manage", "Edit"))
        item = self.table_hospitals.horizontalHeaderItem(6)
        item.setText(_translate("Hospital_Manage", "Delete"))
        self.btn_import.setText(_translate("Hospital_Manage", "Import"))
        self.btn_export.setText(_translate("Hospital_Manage", "Export"))
        self.groupBox.setTitle(_translate("Hospital_Manage", "Search Doctor"))
        self.label_2.setText(_translate("Hospital_Manage", "Name:"))
        self.btn_search.setText(_translate("Hospital_Manage", "Search Hospital"))
        self.tab_search_add.setTabText(self.tab_search_add.indexOf(self.tab_search),
                                       _translate("Hospital_Manage", "Search"))
        self.label_3.setText(_translate("Hospital_Manage", "Name:"))
        self.label_4.setText(_translate("Hospital_Manage", "Phone:"))
        self.label_5.setText(_translate("Hospital_Manage", "Address:"))
        self.label_6.setText(_translate("Hospital_Manage", "Description:"))
        self.btn_add.setText(_translate("Hospital_Manage", "Add"))
        self.tab_search_add.setTabText(self.tab_search_add.indexOf(self.tab_add_new),
                                       _translate("Hospital_Manage", "Add new"))
        self.menuSetting.setTitle(_translate("Hospital_Manage", "Setting"))
        self.menuPrograms.setTitle(_translate("Hospital_Manage", "Programs"))
        self.action_config.setText(_translate("Hospital_Manage", "Config"))
        self.actionPatient.setText(_translate("Hospital_Manage", "Patient"))
        self.actionSchedule.setText(_translate("Hospital_Manage", "Schedule"))
        self.action_doctor.setText(_translate("Hospital_Manage", "Doctor"))
        self.action_patient.setText(_translate("Hospital_Manage", "Patient"))
        self.action_schedule.setText(_translate("Hospital_Manage", "Schedule"))

    def db_get_data_list(self, data_base_setup):
        cnx = mysql.connector.connect(**data_base_setup)
        cursor = cnx.cursor()
        cursor.execute('select * from hospital')
        data_rs = cursor.fetchall()
        return data_rs

    def open_notepad_config(self):
        cm.open_note_pad(cm.DATABASE_CONFIG_PATH)
        self.setupUi()

    def open_doctor_manage(self):
        self.ui_1 = QMainWindow()
        self.ui_1 = Ui_doctor_manage()
        self.ui_1.setupUi()
        self.ui_1.show()
        return self.ui_1

    def open_patient_manage(self):
        self.ui_2 = QMainWindow()
        self.ui_2 = Ui_patient_manage()
        self.ui_2.setupUi()
        self.ui_2.show()

    def open_schedule(self):
        self.ui_3 = QMainWindow()
        self.ui_3 = Ui_schedule_manage()
        self.ui_3.setupUi()
        self.ui_3.show()

    def init_data(self, data):
        for i in range(0, len(data)):
            self.table_hospitals.insertRow(i)
            self.table_hospitals.setItem(i, 0, QtWidgets.QTableWidgetItem(str(data[i][0])))
            self.table_hospitals.setItem(i, 1, QtWidgets.QTableWidgetItem(str(data[i][1])))
            self.table_hospitals.setItem(i, 2, QtWidgets.QTableWidgetItem(str(data[i][2])))
            self.table_hospitals.setItem(i, 3, QtWidgets.QTableWidgetItem(str(data[i][3])))
            self.table_hospitals.setItem(i, 4, QtWidgets.QTableWidgetItem(str(data[i][4])))
            self.table_hospitals.setItem(i, 5, QtWidgets.QTableWidgetItem(str(data[i][4])))
            self.table_hospitals.setItem(i, 6, QtWidgets.QTableWidgetItem(str(data[i][4])))

            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.edit_btn_click)
            self.detail_btn.setText("Edit")
            self.table_hospitals.setCellWidget(i, 5, self.detail_btn)

            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.delete_btn_click)
            self.detail_btn.setText("Delete")
            self.table_hospitals.setCellWidget(i, 6, self.detail_btn)

    def delete_btn_click(self):
        button = QtWidgets.qApp.focusWidget()
        index = self.table_hospitals.indexAt(button.pos())
        table_model = self.table_hospitals.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        ques = QMessageBox.question(self, 'System', f'Are you sure you want to delete hospital: {name}?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ques == QMessageBox.Yes:
            if cm.delete_data(id, 'hospital'):
                QMessageBox.about(self, "System", f"Successfully delete the hospital: {name}")
                self.table_hospitals.clearContents()
                self.table_hospitals.setRowCount(0)
                data_new = cm.select_data('hospital')
                self.init_data(data_new)
            else:
                QMessageBox.about(self, "System", "Delete failed, try again")

    def edit_btn_click(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_form_hospital()
        self.ui.setupUi()
        self.ui.show()

        button = QtWidgets.qApp.focusWidget()
        index = self.table_hospitals.indexAt(button.pos())
        table_model = self.table_hospitals.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        phone_index = table_model.index(index.row(), 2)
        phone = table_model.data(phone_index)
        address_index = table_model.index(index.row(), 3)
        address = table_model.data(address_index)
        description_index = table_model.index(index.row(), 4)
        description = table_model.data(description_index)

        self.ui.txt_id_edit_hospital.setText(id)
        self.ui.txt_name_edit_hospital.setText(name)
        self.ui.txt_phone_edit_hospital.setText(phone)
        self.ui.txt_address_edit_hospital.setText(address)
        self.ui.txt_description_edit_hospital.setText(description)
        self.ui.btn_save_edit_hospital.clicked.connect(self.edit_information_hospital)
        self.ui.btn_cancel_edit_hospital.clicked.connect(self.cancel_form_edit_hospotal)

    def edit_information_hospital(self):
        id = self.ui.txt_id_edit_hospital.text()
        name_new = self.ui.txt_name_edit_hospital.text()
        phone_new = self.ui.txt_phone_edit_hospital.text()
        address_new = self.ui.txt_address_edit_hospital.text()
        description_new = self.ui.txt_description_edit_hospital.text()
        if cm.check_phone(phone_new):
            if cm.edit_data(id, name_new, phone_new, address_new, description_new):
                QMessageBox.information(self, 'Message', f'Successfully edit the hospital: {name_new} ',
                                        QMessageBox.Close)
                self.table_hospitals.clearContents()
                self.table_hospitals.setRowCount(0)
                data_new = cm.select_data('hospital')
                self.init_data(data_new)
                # self.ui = QMainWindow()
                # self.ui = Ui_edit_form_hospital()
                # self.ui.setupUi()
                self.ui.hide()
            else:
                QMessageBox.information(self, 'Message', f'Failed edit the hospital: {name_new} ', QMessageBox.Close)
                self.table_hospitals.clearContents()
                self.table_hospitals.setRowCount(0)
                data_new = cm.select_data('hospital')
                self.init_data(data_new)
                # self.ui = QMainWindow()
                # self.ui = Ui_edit_form_hospital()
                # self.ui.setupUi()
                self.ui.hide()
        else:
            QMessageBox.information(self, 'Message', f'Wrong phone number: {name_new} ', QMessageBox.Close)

    def cancel_form_edit_hospotal(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_form_hospital()
        self.ui.setupUi()
        self.ui.hide()

    def search_hospital(self):
        name_input = self.txt_search.text()
        data = cm.search_data(name_input, 'hospital')
        self.table_hospitals.clearContents()
        self.table_hospitals.setRowCount(0)
        if len(data) != 0:
            self.init_data(data)
        else:
            QMessageBox.information(self, 'Message', 'Nothing found!', QMessageBox.Close)

    def add_new(self):
        name = self.txt_name.text()
        phone_input = self.txt_phone.text()
        address = self.txt_address.text()
        description = self.txt_description.text()
        if cm.check_empty(name):
            if cm.check_phone(phone_input):
                if cm.insert_data(name, phone_input, address, description):
                    QMessageBox.information(self, 'Message', 'Add successfully', QMessageBox.Close)
                    self.txt_name.setText('')
                    self.txt_phone.setText('')
                    self.txt_address.setText('')
                    self.txt_description.setText('')
                    self.table_hospitals.clearContents()
                    self.table_hospitals.setRowCount(0)
                    data_new = cm.select_data('hospital')
                    self.init_data(data_new)
                else:
                    QMessageBox.information(self, 'Message', 'Add fail', QMessageBox.Close)
            else:
                QMessageBox.information(self, 'Message', 'Wrong phone number', QMessageBox.Close)
        else:
            QMessageBox.information(self, 'Message', 'Name is not null', QMessageBox.Close)
    def export_excel(self):
        columnHeaders = []
        for j in range(self.table_hospitals.model().columnCount() - 2):
            columnHeaders.append(self.table_hospitals.horizontalHeaderItem(j).text())
        df = pd.DataFrame(columns=columnHeaders)
        for row in range(self.table_hospitals.rowCount()):
            for col in range(self.table_hospitals.columnCount() - 2):
                df.at[row, columnHeaders[col]] = self.table_hospitals.item(row, col).text()
        wb = Workbook()
        wb = load_workbook('Template/template_hospital_export.xlsx')
        ws1 = wb.worksheets[0]
        offset_row = 1
        offset_col = 0
        row = 1
        for row_data in dataframe_to_rows(df, index=False, header=False):
            col = 1
            for cell_data in row_data:
                ws1.cell(row + offset_row, col + offset_col, cell_data)
                col += 1
            row += 1
        if os.path.exists('Output') == False:
            folder_path = 'Output'
            os.mkdir(folder_path)
        (d, m, y, h, mi, s) = cm.split_date_time()
        wb.save(f'C:\\Users\ADMIN\PycharmProjects\DucBNN_Mock_Project\Output\Hospital_{d}_{m}_{y}_{h}_{mi}_{s}.xlsx')


    def import_excel(self):
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename()
        if 'xlsx' not in file_path:
            QMessageBox.information(self, 'System', 'Please import into excel file!', QMessageBox.Close)
        else:
            wb = load_workbook(file_path)
            ws = wb.worksheets[0]
            list_data = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    list_data.append(cell.value)

            list_data_result = list(cm.divide_chunks(list_data, ws.max_column))
            for i in list_data_result:
                for j in range(0, len(i), ws.max_column):
                    cm.insert_data(i[j], i[j + 1], i[j + 2], i[j + 3])
            self.table_hospitals.clearContents()
            self.table_hospitals.setRowCount(0)
            data_new = cm.select_data('hospital')
            self.init_data(data_new)
            QMessageBox.information(self, 'System', 'Import successfully', QMessageBox.Close)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = UI_hospital_manage()
    ui.setupUi()
    ui.show()
    sys.exit(app.exec_())
