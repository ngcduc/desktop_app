import os
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QMainWindow, QStyledItemDelegate, QTableWidget, QTableWidgetItem
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
import common as cm
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from edit_doctor_patient import Ui_edit_form_doctor_patient

data = cm.select_patient_and_name_hospital()


class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return


class Ui_patient_manage(QMainWindow):
    def __init__(self, parent=None):
        super(Ui_patient_manage, self).__init__(parent)
        self.database_config_obj = None


    def setupUi(self):
        self.setObjectName("patient_manage")
        self.resize(990, 765)
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.label_14 = QtWidgets.QLabel(self.centralwidget)
        self.label_14.setGeometry(QtCore.QRect(400, 0, 121, 51))
        self.setWindowIcon(QtGui.QIcon("./icon/logo.png"))

        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_14.setFont(font)
        self.label_14.setObjectName("label_14")
        self.table_patients = QtWidgets.QTableWidget(self.centralwidget)
        delegate = ReadOnlyDelegate(self)
        self.table_patients.setItemDelegateForColumn(0, delegate)
        self.table_patients.setItemDelegateForColumn(1, delegate)
        self.table_patients.setItemDelegateForColumn(2, delegate)
        self.table_patients.setItemDelegateForColumn(3, delegate)
        self.table_patients.setItemDelegateForColumn(4, delegate)
        self.table_patients.setItemDelegateForColumn(5, delegate)
        self.table_patients.setGeometry(QtCore.QRect(20, 60, 841, 291))
        self.table_patients.setObjectName("table_patients")
        self.table_patients.setColumnCount(8)
        self.table_patients.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_patients.setHorizontalHeaderItem(7, item)
        self.btn_import_patient = QtWidgets.QPushButton(self.centralwidget)
        self.btn_import_patient.clicked.connect(self.import_excel)
        self.btn_import_patient.setGeometry(QtCore.QRect(870, 70, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_import_patient.setFont(font)
        self.btn_import_patient.setObjectName("btn_import_patient")
        self.btn_export_patient = QtWidgets.QPushButton(self.centralwidget)
        self.btn_export_patient.clicked.connect(self.export_excel)
        self.btn_export_patient.setGeometry(QtCore.QRect(870, 120, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_export_patient.setFont(font)
        self.btn_export_patient.setObjectName("btn_export_patient")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(20, 360, 941, 341))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.groupBox = QtWidgets.QGroupBox(self.tab)
        self.groupBox.setGeometry(QtCore.QRect(20, 20, 861, 171))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.txt_patient = QtWidgets.QLineEdit(self.groupBox)
        self.txt_patient.setGeometry(QtCore.QRect(130, 31, 691, 31))
        self.txt_patient.setObjectName("txt_patient")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 40, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.search_patient = QtWidgets.QPushButton(self.groupBox)
        self.search_patient.setGeometry(QtCore.QRect(370, 90, 121, 41))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.search_patient.setFont(font)
        self.search_patient.setObjectName("search_patient")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.label_2 = QtWidgets.QLabel(self.tab_2)
        self.label_2.setGeometry(QtCore.QRect(40, 30, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.tab_2)
        self.label_3.setGeometry(QtCore.QRect(40, 70, 61, 31))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.tab_2)
        self.label_4.setGeometry(QtCore.QRect(40, 120, 61, 31))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.tab_2)
        self.label_5.setGeometry(QtCore.QRect(40, 170, 81, 31))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.tab_2)
        self.label_6.setGeometry(QtCore.QRect(40, 220, 101, 31))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.txt_name_patient = QtWidgets.QLineEdit(self.tab_2)
        self.txt_name_patient.setGeometry(QtCore.QRect(150, 21, 721, 31))
        self.txt_name_patient.setObjectName("txt_name_patient")
        self.txt_phone_patient = QtWidgets.QLineEdit(self.tab_2)
        self.txt_phone_patient.setGeometry(QtCore.QRect(150, 70, 721, 31))
        self.txt_phone_patient.setObjectName("txt_phone_patient")
        self.txt_emai_patient = QtWidgets.QLineEdit(self.tab_2)
        self.txt_emai_patient.setGeometry(QtCore.QRect(150, 120, 721, 31))
        self.txt_emai_patient.setObjectName("txt_emai_patient")
        self.txt_address_patient = QtWidgets.QLineEdit(self.tab_2)
        self.txt_address_patient.setGeometry(QtCore.QRect(150, 170, 721, 31))
        self.txt_address_patient.setObjectName("txt_address_patient")
        self.combo_hospital_patient = QtWidgets.QComboBox(self.tab_2)
        wordlist = cm.select_table_name('hospital').tolist()
        wordList_1 = []
        for i in wordlist:
            t = ' '.join(i)
            wordList_1.append(t)
        self.combo_hospital_patient.addItems(wordList_1)
        self.combo_hospital_patient.setEditable(True)
        self.combo_hospital_patient.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.combo_hospital_patient.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.combo_hospital_patient.setGeometry(QtCore.QRect(150, 221, 251, 31))
        self.combo_hospital_patient.setObjectName("combo_hospital_patient")

        self.btn_add_patient = QtWidgets.QPushButton(self.tab_2)
        self.btn_add_patient.clicked.connect(self.add_patient)
        self.btn_add_patient.setGeometry(QtCore.QRect(400, 260, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btn_add_patient.setFont(font)
        self.btn_add_patient.setObjectName("btn_add_patient")
        self.tabWidget.addTab(self.tab_2, "")
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 979, 25))
        self.menubar.setObjectName("menubar")
        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.retranslateUi()
        self.tabWidget.setCurrentIndex(0)
        self.database_config_obj = cm.load_config(cm.DATABASE_CONFIG_PATH)
        self.init_data(data)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("patient_manage", "Patients Manage"))
        self.label_14.setText(_translate("patient_manage", "Patients"))
        item = self.table_patients.horizontalHeaderItem(0)
        item.setText(_translate("patient_manage", "ID"))
        item = self.table_patients.horizontalHeaderItem(1)
        item.setText(_translate("patient_manage", "Name"))
        item = self.table_patients.horizontalHeaderItem(2)
        item.setText(_translate("patient_manage", "Phone"))
        item = self.table_patients.horizontalHeaderItem(3)
        item.setText(_translate("patient_manage", "Email"))
        item = self.table_patients.horizontalHeaderItem(4)
        item.setText(_translate("patient_manage", "Address"))
        item = self.table_patients.horizontalHeaderItem(5)
        item.setText(_translate("patient_manage", "Hospital"))
        item = self.table_patients.horizontalHeaderItem(6)
        item.setText(_translate("patient_manage", "Edit"))
        item = self.table_patients.horizontalHeaderItem(7)
        item.setText(_translate("patient_manage", "Delete"))
        self.btn_import_patient.setText(_translate("patient_manage", "Import"))
        self.btn_export_patient.setText(_translate("patient_manage", "Export"))
        self.groupBox.setTitle(_translate("patient_manage", "Search Patient"))
        self.label.setText(_translate("patient_manage", "Name:"))
        self.search_patient.setText(_translate("patient_manage", "Search Patient"))
        self.search_patient.clicked.connect(self.func_search_patient)
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("patient_manage", "Search"))
        self.label_2.setText(_translate("patient_manage", "Name:"))
        self.label_3.setText(_translate("patient_manage", "Phone:"))
        self.label_4.setText(_translate("patient_manage", "Email:"))
        self.label_5.setText(_translate("patient_manage", "Address:"))
        self.label_6.setText(_translate("patient_manage", "Hostipal id:"))
        self.btn_add_patient.setText(_translate("patient_manage", "Add"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("patient_manage", "Add new"))

    def init_data(self, data):
        for i in range(0, len(data)):
            self.table_patients.insertRow(i)
            self.table_patients.setItem(i, 0, QtWidgets.QTableWidgetItem(str(data[i][0])))
            self.table_patients.setItem(i, 1, QtWidgets.QTableWidgetItem(str(data[i][1])))
            self.table_patients.setItem(i, 2, QtWidgets.QTableWidgetItem(str(data[i][2])))
            self.table_patients.setItem(i, 3, QtWidgets.QTableWidgetItem(str(data[i][3])))
            self.table_patients.setItem(i, 4, QtWidgets.QTableWidgetItem(str(data[i][4])))
            self.table_patients.setItem(i, 5, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.table_patients.setItem(i, 6, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.table_patients.setItem(i, 7, QtWidgets.QTableWidgetItem(str(data[i][5])))

            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.edit_btn_click)
            self.detail_btn.setText("Edit")
            self.table_patients.setCellWidget(i, 6, self.detail_btn)

            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.delete_btn_click)
            self.detail_btn.setText("Delete")
            self.table_patients.setCellWidget(i, 7, self.detail_btn)

    def edit_btn_click(self):

        self.ui = QMainWindow()
        self.ui = Ui_edit_form_doctor_patient()
        self.ui.setupUi()
        self.ui.show()
        button = QtWidgets.qApp.focusWidget()
        index = self.table_patients.indexAt(button.pos())
        table_model = self.table_patients.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        phone_index = table_model.index(index.row(), 2)
        phone = table_model.data(phone_index)
        email_index = table_model.index(index.row(), 3)
        email = table_model.data(email_index)
        address_index = table_model.index(index.row(), 4)
        address = table_model.data(address_index)
        hospital_name_index = table_model.index(index.row(), 5)
        hospital_name = table_model.data(hospital_name_index)
        hospital_id = int(cm.convert_name_to_id(hospital_name, 'hospital'))

        self.ui.txt_id_edit_doctor_patient.setText(id)
        self.ui.txt_name_edit_doctor_patient.setText(name)
        self.ui.txt_phone_edit_doctor_patient.setText(phone)
        self.ui.txt_email_edit_doctor_patient.setText(email)
        self.ui.txt_address_edit_doctor_patient.setText(address)
        self.ui.combo_edit_doctor_patient.setCurrentText(hospital_name)

        self.ui.btn_save_edit_doctor_patient.clicked.connect(self.edit_patient)
        self.ui.btn_cancel_edit_doctor_patient.clicked.connect(self.cancel_edit_patient)
    def cancel_edit_patient(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_form_doctor_patient()
        self.ui.setupUi()
        self.ui.hide()

    def edit_patient(self):
        id = self.ui.txt_id_edit_doctor_patient.text()
        name_new = self.ui.txt_name_edit_doctor_patient.text()
        phone_new = self.ui.txt_phone_edit_doctor_patient.text()
        email_new = self.ui.txt_email_edit_doctor_patient.text()
        address_new = self.ui.txt_address_edit_doctor_patient.text()
        hospital_id = int(cm.convert_name_to_id(self.ui.combo_edit_doctor_patient.currentText(), 'hospital'))
        if cm.check_phone(phone_new):
            if cm.check_email(email_new):
                if cm.edit_data_doctor(id, name_new, phone_new, email_new, address_new, hospital_id, 'patient'):
                    QMessageBox.information(self, 'Message', f'Successfully edit the patient: {name_new} ', QMessageBox.Close)
                    self.table_patients.clearContents()
                    self.table_patients.setRowCount(0)
                    data_new = cm.select_patient_and_name_hospital()
                    self.init_data(data_new)
                    self.ui = QMainWindow()
                    self.ui = Ui_edit_form_doctor_patient()
                    self.ui.setupUi()
                    self.ui.hide()
                else:
                    QMessageBox.information(self, 'Message', f'Failed edit the patient: {name_new} ', QMessageBox.Close)
                    self.table_patients.clearContents()
                    self.table_patients.setRowCount(0)
                    data_new = cm.select_patient_and_name_hospital()
                    self.init_data(data_new)
                    self.ui = QMainWindow()
                    self.ui = Ui_edit_form_doctor_patient()
                    self.ui.setupUi()
                    self.ui.hide()
            else:
                QMessageBox.information(self, 'System', 'Wrong email number', QMessageBox.Close)

        else:
            QMessageBox.information(self, 'System', 'Wrong phone number', QMessageBox.Close)

    def delete_btn_click(self):
        button = QtWidgets.qApp.focusWidget()
        index = self.table_patients.indexAt(button.pos())
        table_model = self.table_patients.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        ques = QMessageBox.question(self, 'System', f'Are you sure you want to delete hospital: {name}?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ques == QMessageBox.Yes:
            if cm.delete_data(id, 'patient'):
                QMessageBox.about(self, "System", f"Successfully delete the patient: {name}")
                self.table_patients.clearContents()
                self.table_patients.setRowCount(0)
                data_new = cm.select_patient_and_name_hospital()
                self.init_data(data_new)
            else:
                QMessageBox.about(self, "System", "Delete failed, try again")

    def func_search_patient(self):
        name_input = self.txt_patient.text()
        data = cm.search_data(name_input, 'patient')
        self.table_patients.clearContents()
        self.table_patients.setRowCount(0)
        if len(data) != 0:
            self.init_data(data)
        else:
            QMessageBox.information(self, 'Message', 'Nothing found!', QMessageBox.Close)

    def add_patient(self):
        name = self.txt_name_patient.text()
        phone = self.txt_phone_patient.text()
        email = self.txt_emai_patient.text()
        address = self.txt_address_patient.text()
        hospital_name = self.combo_hospital_patient.currentText()
        hospital_id= int(cm.convert_name_to_id(hospital_name,'hospital'))
        if cm.check_empty(name):
            if cm.check_phone(phone):
                if cm.check_email(email):
                    if cm.insert_doctor(name, phone, email, address, hospital_id, 'patient'):
                        QMessageBox.information(self, 'System', 'Add successfully', QMessageBox.Close)
                        self.table_patients.clearContents()
                        self.table_patients.setRowCount(0)
                        data_new = cm.select_patient_and_name_hospital()
                        self.init_data(data_new)
                    else:
                        QMessageBox.information(self, 'Message', 'Add fail', QMessageBox.Close)
                        self.table_patients.clearContents()
                        self.table_patients.setRowCount(0)
                        data_new = cm.select_patient_and_name_hospital()
                        self.init_data(data_new)
                else:
                    QMessageBox.information(self, 'System', 'Wrong email number', QMessageBox.Close)

            else:
                QMessageBox.information(self, 'System', 'Wrong phone number', QMessageBox.Close)
        else:
            QMessageBox.information(self, 'Message', 'Name is not null', QMessageBox.Close)


    def import_excel(self):

        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename()
        if 'xlsx' not in file_path:
            QMessageBox.information(self, 'System', 'Please import into excel file!', QMessageBox.Close)
        else:
            wb = load_workbook(file_path)
            ws = wb.worksheets[0]
            list_data = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    list_data.append(cell.value)
            list_data_result = list(cm.divide_chunks(list_data, ws.max_column))

            for i in list_data_result:
                for j in range(0, len(i), ws.max_column):
                    cm.insert_doctor(i[j], i[j + 1], i[j + 2], i[j + 3], i[j + 4],'patient')
            self.table_patients.clearContents()
            self.table_patients.setRowCount(0)
            data_new = cm.select_patient_and_name_hospital()
            self.init_data(data_new)
            QMessageBox.information(self, 'Message', 'Import successfully', QMessageBox.Close)

    def export_excel(self):
        columnHeaders = []
        for j in range(self.table_patients.model().columnCount() - 2):
            columnHeaders.append(self.table_patients.horizontalHeaderItem(j).text())
        df = pd.DataFrame(columns=columnHeaders)
        for row in range(self.table_patients.rowCount()):
            for col in range(self.table_patients.columnCount() - 2):
                df.at[row, columnHeaders[col]] = self.table_patients.item(row, col).text()
        wb = Workbook()
        wb = load_workbook('Template/template_patient_export.xlsx')
        ws1 = wb.worksheets[0]
        offset_row = 1
        offset_col = 0
        row = 1
        for row_data in dataframe_to_rows(df, index=False, header=False):
            col = 1
            for cell_data in row_data:
                ws1.cell(row + offset_row, col + offset_col, cell_data)
                col += 1
            row += 1
        if os.path.exists('Output') == False:
            folder_path = 'Output'
            os.mkdir(folder_path)
        (d, m, y, h, mi, s) = cm.split_date_time()
        wb.save(f'C:\\Users\ADMIN\PycharmProjects\DucBNN_Mock_Project\Output\Patient_{d}_{m}_{y}_{h}_{mi}_{s}.xlsx')


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = Ui_patient_manage()
    ui.setupUi()
    ui.show()
    sys.exit(app.exec_())
