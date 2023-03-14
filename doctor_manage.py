import os
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QMainWindow, QStyledItemDelegate
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
import common as cm
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from edit_doctor_patient import Ui_edit_form_doctor_patient

data = cm.select_doctor_and_name_hospital()


class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return


class Ui_doctor_manage(QMainWindow):
    def __init__(self, parent=None):
        super(Ui_doctor_manage, self).__init__(parent)
        self.database_config_obj = None


    def setupUi(self):
        self.setObjectName("doctor_manage")
        self.resize(990, 769)
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(420, 10, 121, 51))
        self.setWindowIcon(QtGui.QIcon("./icon/logo.png"))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.table_doctor = QtWidgets.QTableWidget(self.centralwidget)
        delegate = ReadOnlyDelegate(self)
        self.table_doctor.setItemDelegateForColumn(0, delegate)
        self.table_doctor.setItemDelegateForColumn(1, delegate)
        self.table_doctor.setItemDelegateForColumn(2, delegate)
        self.table_doctor.setItemDelegateForColumn(3, delegate)
        self.table_doctor.setItemDelegateForColumn(4, delegate)
        self.table_doctor.setItemDelegateForColumn(5, delegate)
        self.table_doctor.setGeometry(QtCore.QRect(20, 70, 831, 281))
        self.table_doctor.setObjectName("table_doctor")
        self.table_doctor.setColumnCount(8)
        self.table_doctor.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_doctor.setHorizontalHeaderItem(7, item)
        self.btn_import_doctor = QtWidgets.QPushButton(self.centralwidget)
        self.btn_import_doctor.clicked.connect(self.import_doctor)
        self.btn_import_doctor.setGeometry(QtCore.QRect(870, 90, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_import_doctor.setFont(font)
        self.btn_import_doctor.setObjectName("btn_import_doctor")
        self.btn_export_doctor = QtWidgets.QPushButton(self.centralwidget)
        self.btn_export_doctor.clicked.connect(self.export_doctor)
        self.btn_export_doctor.setGeometry(QtCore.QRect(870, 150, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_export_doctor.setFont(font)
        self.btn_export_doctor.setObjectName("btn_export_doctor")
        self.tab_search_add_doctor = QtWidgets.QTabWidget(self.centralwidget)
        self.tab_search_add_doctor.setGeometry(QtCore.QRect(20, 370, 941, 351))
        self.tab_search_add_doctor.setObjectName("tab_search_add_doctor")
        self.tab_search_doctor = QtWidgets.QWidget()
        self.tab_search_doctor.setObjectName("tab_search_doctor")
        self.groupBox_2 = QtWidgets.QGroupBox(self.tab_search_doctor)
        self.groupBox_2.setGeometry(QtCore.QRect(20, 10, 811, 141))
        self.groupBox_2.setObjectName("groupBox_2")
        self.label_8 = QtWidgets.QLabel(self.groupBox_2)
        self.label_8.setGeometry(QtCore.QRect(40, 40, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.txt_search_doctor = QtWidgets.QLineEdit(self.groupBox_2)
        self.txt_search_doctor.setGeometry(QtCore.QRect(130, 30, 661, 31))
        self.txt_search_doctor.setObjectName("txt_search_doctor")
        self.btn_search_doctor = QtWidgets.QPushButton(self.groupBox_2)
        self.btn_search_doctor.clicked.connect(self.search_doctor)
        self.btn_search_doctor.setGeometry(QtCore.QRect(330, 90, 151, 31))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.btn_search_doctor.setFont(font)
        self.btn_search_doctor.setObjectName("btn_search_doctor")
        self.tab_search_add_doctor.addTab(self.tab_search_doctor, "")
        self.tab_add_new_doctor = QtWidgets.QWidget()
        self.tab_add_new_doctor.setObjectName("tab_add_new_doctor")
        self.label_9 = QtWidgets.QLabel(self.tab_add_new_doctor)
        self.label_9.setGeometry(QtCore.QRect(40, 40, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(self.tab_add_new_doctor)
        self.label_10.setGeometry(QtCore.QRect(40, 90, 55, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(self.tab_add_new_doctor)
        self.label_11.setGeometry(QtCore.QRect(40, 140, 71, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.tab_add_new_doctor)
        self.label_12.setGeometry(QtCore.QRect(40, 190, 101, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_12.setFont(font)
        self.label_12.setObjectName("label_12")
        self.txt_name_doctor = QtWidgets.QLineEdit(self.tab_add_new_doctor)
        self.txt_name_doctor.setGeometry(QtCore.QRect(180, 31, 621, 31))
        self.txt_name_doctor.setObjectName("txt_name_doctor")
        self.txt_phone_doctor = QtWidgets.QLineEdit(self.tab_add_new_doctor)
        self.txt_phone_doctor.setGeometry(QtCore.QRect(180, 80, 621, 31))
        self.txt_phone_doctor.setObjectName("txt_phone_doctor")
        self.txt_email_doctor = QtWidgets.QLineEdit(self.tab_add_new_doctor)
        self.txt_email_doctor.setGeometry(QtCore.QRect(180, 130, 621, 31))
        self.txt_email_doctor.setObjectName("txt_email_doctor")
        self.txt_address_doctor = QtWidgets.QLineEdit(self.tab_add_new_doctor)
        self.txt_address_doctor.setGeometry(QtCore.QRect(180, 180, 621, 31))
        self.txt_address_doctor.setObjectName("txt_address_doctor")
        self.btn_add_doctor = QtWidgets.QPushButton(self.tab_add_new_doctor)
        self.btn_add_doctor.clicked.connect(self.add_new_doctor)
        self.btn_add_doctor.setGeometry(QtCore.QRect(400, 280, 89, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btn_add_doctor.setFont(font)
        self.btn_add_doctor.setObjectName("btn_add_doctor")
        self.label_13 = QtWidgets.QLabel(self.tab_add_new_doctor)
        self.label_13.setGeometry(QtCore.QRect(40, 240, 101, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_13.setFont(font)
        self.label_13.setObjectName("label_13")
        self.combo_hospital_doctor = QtWidgets.QComboBox(self.tab_add_new_doctor)

        wordlist = cm.select_table_name('hospital').tolist()
        wordList_1 = []
        for i in wordlist:
            t = ' '.join(i)
            wordList_1.append(t)
        self.combo_hospital_doctor.addItems(wordList_1)
        self.combo_hospital_doctor.setEditable(True)
        self.combo_hospital_doctor.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.combo_hospital_doctor.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.combo_hospital_doctor.setGeometry(QtCore.QRect(180, 231, 201, 31))
        self.combo_hospital_doctor.setObjectName("combo_hospital_doctor")
        self.tab_search_add_doctor.addTab(self.tab_add_new_doctor, "")
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 981, 25))
        self.menubar.setObjectName("menubar")
        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.retranslateUi()
        self.tab_search_add_doctor.setCurrentIndex(1)
        self.init_data(data)
        self.database_config_obj = cm.load_config(cm.DATABASE_CONFIG_PATH)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("doctor_manage", "Doctor Manage"))
        self.label_7.setText(_translate("doctor_manage", "Doctors"))
        item = self.table_doctor.horizontalHeaderItem(0)
        item.setText(_translate("doctor_manage", "ID"))
        item = self.table_doctor.horizontalHeaderItem(1)
        item.setText(_translate("doctor_manage", "Name"))
        item = self.table_doctor.horizontalHeaderItem(2)
        item.setText(_translate("doctor_manage", "Phone"))
        item = self.table_doctor.horizontalHeaderItem(3)
        item.setText(_translate("doctor_manage", "Email"))
        item = self.table_doctor.horizontalHeaderItem(4)
        item.setText(_translate("doctor_manage", "Adress"))
        item = self.table_doctor.horizontalHeaderItem(5)
        item.setText(_translate("doctor_manage", "Hospital"))
        item = self.table_doctor.horizontalHeaderItem(6)
        item.setText(_translate("doctor_manage", "Edit"))
        item = self.table_doctor.horizontalHeaderItem(7)
        item.setText(_translate("doctor_manage", "Delete"))
        self.btn_import_doctor.setText(_translate("doctor_manage", "Import"))
        self.btn_export_doctor.setText(_translate("doctor_manage", "Export"))
        self.groupBox_2.setTitle(_translate("doctor_manage", "Search Doctor"))
        self.label_8.setText(_translate("doctor_manage", "Name:"))
        self.btn_search_doctor.setText(_translate("doctor_manage", "Search Doctor"))
        self.tab_search_add_doctor.setTabText(self.tab_search_add_doctor.indexOf(self.tab_search_doctor),
                                              _translate("doctor_manage", "Search"))
        self.label_9.setText(_translate("doctor_manage", "Name:"))
        self.label_10.setText(_translate("doctor_manage", "Phone:"))
        self.label_11.setText(_translate("doctor_manage", "Email:"))
        self.label_12.setText(_translate("doctor_manage", "Address:"))
        self.btn_add_doctor.setText(_translate("doctor_manage", "Add"))
        self.label_13.setText(_translate("doctor_manage", "Hospital id:"))
        self.tab_search_add_doctor.setTabText(self.tab_search_add_doctor.indexOf(self.tab_add_new_doctor),
                                              _translate("doctor_manage", "Add new"))

    def init_data(self, data):
        for i in range(0, len(data)):
            self.table_doctor.insertRow(i)
            self.table_doctor.setItem(i, 0, QtWidgets.QTableWidgetItem(str(data[i][0])))
            self.table_doctor.setItem(i, 1, QtWidgets.QTableWidgetItem(str(data[i][1])))
            self.table_doctor.setItem(i, 2, QtWidgets.QTableWidgetItem(str(data[i][2])))
            self.table_doctor.setItem(i, 3, QtWidgets.QTableWidgetItem(str(data[i][3])))
            self.table_doctor.setItem(i, 4, QtWidgets.QTableWidgetItem(str(data[i][4])))
            self.table_doctor.setItem(i, 5, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.table_doctor.setItem(i, 6, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.table_doctor.setItem(i, 7, QtWidgets.QTableWidgetItem(str(data[i][5])))

            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.edit_btn_click)
            self.detail_btn.setText("Edit")
            self.table_doctor.setCellWidget(i, 6, self.detail_btn)

            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.delete_btn_click)
            self.detail_btn.setText("Delete")
            self.table_doctor.setCellWidget(i, 7, self.detail_btn)

    def edit_btn_click(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_form_doctor_patient()
        self.ui.setupUi()
        self.ui.show()
        button = QtWidgets.qApp.focusWidget()
        index = self.table_doctor.indexAt(button.pos())
        table_model = self.table_doctor.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        phone_index = table_model.index(index.row(), 2)
        phone = table_model.data(phone_index)
        email_index = table_model.index(index.row(), 3)
        email = table_model.data(email_index)
        address_index = table_model.index(index.row(), 4)
        address = table_model.data(address_index)
        hospital_name_index = table_model.index(index.row(), 5)
        hospital_name = table_model.data(hospital_name_index)
        hospital_id = int(cm.convert_name_to_id(hospital_name, 'hospital'))

        self.ui.txt_id_edit_doctor_patient.setText(id)
        self.ui.txt_name_edit_doctor_patient.setText(name)
        self.ui.txt_phone_edit_doctor_patient.setText(phone)
        self.ui.txt_email_edit_doctor_patient.setText(email)
        self.ui.txt_address_edit_doctor_patient.setText(address)
        self.ui.combo_edit_doctor_patient.setCurrentText(hospital_name)

        self.ui.btn_save_edit_doctor_patient.clicked.connect(self.edit_doctor)
        self.ui.btn_cancel_edit_doctor_patient.clicked.connect(self.cancel_edit_doctor)

    def cancel_edit_doctor(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_form_doctor_patient()
        self.ui.setupUi()
        self.ui.hide()

    def edit_doctor(self):
        id = self.ui.txt_id_edit_doctor_patient.text()
        name_new = self.ui.txt_name_edit_doctor_patient.text()
        phone_new = self.ui.txt_phone_edit_doctor_patient.text()
        email_new = self.ui.txt_email_edit_doctor_patient.text()
        address_new = self.ui.txt_address_edit_doctor_patient.text()
        hospital_id = int(cm.convert_name_to_id(self.ui.combo_edit_doctor_patient.currentText(), 'hospital'))
        if cm.check_phone(phone_new):
            if cm.check_email(email_new):
                if cm.edit_data_doctor(id, name_new, phone_new, email_new, address_new, hospital_id, 'doctor'):
                    QMessageBox.information(self, 'Message', f'Successfully edit the doctor: {name_new} ',
                                            QMessageBox.Close)
                    self.table_doctor.clearContents()
                    self.table_doctor.setRowCount(0)
                    data_new = cm.select_doctor_and_name_hospital()
                    self.init_data(data_new)
                    self.ui = QMainWindow()
                    self.ui = Ui_edit_form_doctor_patient()
                    self.ui.setupUi()
                    self.ui.hide()
                else:
                    QMessageBox.information(self, 'Message', f'Failed edit the doctor: {name_new} ', QMessageBox.Close)
                    self.table_doctor.clearContents()
                    self.table_doctor.setRowCount(0)
                    data_new = cm.select_doctor_and_name_hospital()
                    self.init_data(data_new)
                    self.ui = QMainWindow()
                    self.ui = Ui_edit_form_doctor_patient()
                    self.ui.setupUi()
                    self.ui.hide()
            else:
                QMessageBox.information(self, 'Message', 'Wrong email format!', QMessageBox.Close)
        else:
            QMessageBox.information(self, 'Message', 'Wrong phone number!', QMessageBox.Close)

    def delete_btn_click(self):
        button = QtWidgets.qApp.focusWidget()
        index = self.table_doctor.indexAt(button.pos())
        table_model = self.table_doctor.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        ques = QMessageBox.question(self, 'System', f'Are you sure you want to delete doctor: {name}?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ques == QMessageBox.Yes:
            if cm.delete_data(id, 'doctor'):
                QMessageBox.about(self, "System", f"Successfully delete the doctor: {name}")
                self.table_doctor.clearContents()
                self.table_doctor.setRowCount(0)
                data_new = cm.select_doctor_and_name_hospital()
                self.init_data(data_new)
            else:
                QMessageBox.about(self, "System", "Delete failed, try again!")

    def search_doctor(self):
        name_input = self.txt_search_doctor.text()
        data = cm.search_data(name_input, 'doctor')
        self.table_doctor.clearContents()
        self.table_doctor.setRowCount(0)
        if len(data) != 0:
            self.init_data(data)
        else:
            QMessageBox.information(self, 'Message', 'Nothing found!', QMessageBox.Close)

    def add_new_doctor(self):
        name = self.txt_name_doctor.text()
        phone = self.txt_phone_doctor.text()
        email = self.txt_email_doctor.text()
        address = self.txt_address_doctor.text()
        hospital_name = self.combo_hospital_doctor.currentText()
        hospital_id = int(cm.convert_name_to_id(hospital_name, 'hospital'))
        if cm.check_empty(name):
            if cm.check_phone(phone):
                if cm.check_email(email):
                    if cm.insert_doctor(name, phone, email, address, hospital_id, 'doctor'):
                        QMessageBox.information(self, 'System', 'Add successfully', QMessageBox.Close)
                        self.txt_name_doctor.setText('')
                        self.txt_phone_doctor.setText('')
                        self.txt_email_doctor.setText('')
                        self.txt_address_doctor.setText('')

                        self.table_doctor.clearContents()
                        self.table_doctor.setRowCount(0)
                        data_new = cm.select_doctor_and_name_hospital()
                        self.init_data(data_new)
                    else:
                        QMessageBox.information(self, 'Message', 'Add fail', QMessageBox.Close)
                        self.table_doctor.clearContents()
                        self.table_doctor.setRowCount(0)
                        data_new = cm.select_doctor_and_name_hospital()
                        self.init_data(data_new)
                else:
                    QMessageBox.information(self, 'System', 'Wrong email format!', QMessageBox.Close)

            else:
                QMessageBox.information(self, 'System', 'Wrong phone number format!', QMessageBox.Close)
        else:
            QMessageBox.information(self, 'Message', 'Name is not null', QMessageBox.Close)


    def import_doctor(self):
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename()
        if 'xlsx' not in file_path:
            QMessageBox.information(self, 'System', 'Please import into excel file!', QMessageBox.Close)
        else:
            wb = load_workbook(file_path)
            ws = wb.worksheets[0]
            worksheet_name = wb.sheetnames

            list_a = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    list_a.append(cell.value)
            list_result = list(cm.divide_chunks(list_a, ws.max_column))

            for i in list_result:
                for j in range(0, len(i), ws.max_column):
                    cm.insert_doctor(i[j], i[j + 1], i[j + 2], i[j + 3], i[j + 4], 'doctor')
            self.table_doctor.clearContents()
            self.table_doctor.setRowCount(0)
            data_new = cm.select_doctor_and_name_hospital()
            self.init_data(data_new)
            QMessageBox.information(self, 'System', 'Import successfully', QMessageBox.Close)

    def export_doctor(self):
        columnHeaders = []
        for j in range(self.table_doctor.model().columnCount() - 2):
            columnHeaders.append(self.table_doctor.horizontalHeaderItem(j).text())

        df = pd.DataFrame(columns=columnHeaders)
        for row in range(self.table_doctor.rowCount()):
            for col in range(self.table_doctor.columnCount() - 2):
                df.at[row, columnHeaders[col]] = self.table_doctor.item(row, col).text()
        wb = Workbook()
        wb = load_workbook('Template/template_doctor_export.xlsx')
        ws1 = wb.worksheets[0]
        offset_row = 1
        offset_col = 0
        row = 1
        for row_data in dataframe_to_rows(df, index=False, header=False):
            col = 1
            for cell_data in row_data:
                ws1.cell(row + offset_row, col + offset_col, cell_data)
                col += 1
            row += 1
        # if os.path.exists('Output'):
        if os.path.exists('Output') == False:
            folder_path = 'Output'
            os.mkdir(folder_path)
        (d, m, y, h, mi, s) = cm.split_date_time()
        wb.save(f'C:\\Users\ADMIN\PycharmProjects\DucBNN_Mock_Project\Output\Doctor{d}_{m}_{y}_{h}_{mi}_{s}.xlsx')


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = Ui_doctor_manage()
    ui.setupUi()
    ui.show()
    sys.exit(app.exec_())
