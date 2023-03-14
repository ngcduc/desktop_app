import os

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QMainWindow, QStyledItemDelegate
import sys

from openpyxl.utils.dataframe import dataframe_to_rows

import common as cm
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
from edit_schedule import Ui_edit_schedule
from datetime import datetime
import datetime

data = cm.select_schedule()


class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return


class Ui_schedule_manage(QMainWindow):
    def __init__(self, parent=None):
        super(Ui_schedule_manage, self).__init__(parent)
        self.database_config_obj = None


    def setupUi(self):
        self.setObjectName("schedule_manage")
        self.resize(986, 768)
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.label_14 = QtWidgets.QLabel(self.centralwidget)
        self.label_14.setGeometry(QtCore.QRect(430, 0, 131, 51))
        self.setWindowIcon(QtGui.QIcon("./icon/logo.png"))

        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_14.setFont(font)
        self.label_14.setObjectName("label_14")
        self.table_schedule = QtWidgets.QTableWidget(self.centralwidget)
        delegate = ReadOnlyDelegate(self)
        self.table_schedule.setItemDelegateForColumn(0, delegate)
        self.table_schedule.setItemDelegateForColumn(1, delegate)
        self.table_schedule.setItemDelegateForColumn(2, delegate)
        self.table_schedule.setItemDelegateForColumn(4, delegate)
        self.table_schedule.setItemDelegateForColumn(5, delegate)
        self.table_schedule.setItemDelegateForColumn(3, delegate)
        self.table_schedule.setGeometry(QtCore.QRect(20, 60, 841, 291))
        self.table_schedule.setObjectName("table_schedule")
        self.table_schedule.setColumnCount(8)
        self.table_schedule.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_schedule.setHorizontalHeaderItem(7, item)
        self.btn_import_schedule = QtWidgets.QPushButton(self.centralwidget)
        self.btn_import_schedule.clicked.connect(self.import_schedule)
        self.btn_import_schedule.setGeometry(QtCore.QRect(870, 80, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_import_schedule.setFont(font)
        self.btn_import_schedule.setObjectName("btn_import_schedule")
        self.btn_export_schedule = QtWidgets.QPushButton(self.centralwidget)
        self.btn_export_schedule.clicked.connect(self.export_schedule)
        self.btn_export_schedule.setGeometry(QtCore.QRect(870, 130, 101, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_export_schedule.setFont(font)
        self.btn_export_schedule.setObjectName("btn_export_schedule")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(20, 380, 941, 331))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.groupBox = QtWidgets.QGroupBox(self.tab)
        self.groupBox.setGeometry(QtCore.QRect(30, 30, 851, 211))
        self.groupBox.setObjectName("groupBox")
        self.label_6 = QtWidgets.QLabel(self.groupBox)
        self.label_6.setGeometry(QtCore.QRect(40, 40, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.txt_search_schedule = QtWidgets.QLineEdit(self.groupBox)
        self.txt_search_schedule.setGeometry(QtCore.QRect(140, 40, 651, 31))
        self.txt_search_schedule.setObjectName("txt_search_schedule")
        self.label_7 = QtWidgets.QLabel(self.groupBox)
        self.label_7.setGeometry(QtCore.QRect(40, 100, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self.groupBox)
        self.label_8.setGeometry(QtCore.QRect(520, 100, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.data_from_search_schedule = QtWidgets.QDateTimeEdit(self.groupBox)
        self.data_from_search_schedule.setGeometry(QtCore.QRect(140, 91, 194, 31))
        self.data_from_search_schedule.setObjectName("data_from_search_schedule")
        self.date_to_search_schedule = QtWidgets.QDateTimeEdit(self.groupBox)
        self.date_to_search_schedule.setDate(datetime.datetime.now())
        self.date_to_search_schedule.setGeometry(QtCore.QRect(600, 90, 194, 31))
        self.date_to_search_schedule.setObjectName("date_to_search_schedule")
        self.btn_search_schdule = QtWidgets.QPushButton(self.groupBox)
        self.btn_search_schdule.clicked.connect(self.search_schedule)
        self.btn_search_schdule.setGeometry(QtCore.QRect(350, 160, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.btn_search_schdule.setFont(font)
        self.btn_search_schdule.setObjectName("btn_search_schdule")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.label_2 = QtWidgets.QLabel(self.tab_2)
        self.label_2.setGeometry(QtCore.QRect(50, 30, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.tab_2)
        self.label_3.setGeometry(QtCore.QRect(50, 80, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.tab_2)
        self.label_4.setGeometry(QtCore.QRect(50, 130, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.tab_2)
        self.label_5.setGeometry(QtCore.QRect(50, 180, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.txt_name_schedule = QtWidgets.QLineEdit(self.tab_2)
        self.txt_name_schedule.setGeometry(QtCore.QRect(170, 20, 721, 31))
        self.txt_name_schedule.setObjectName("txt_name_schedule")
        self.combo_doctor_schedule = QtWidgets.QComboBox(self.tab_2)
        wordlist = cm.select_table_name('doctor').tolist()
        wordList_1 = []
        for i in wordlist:
            t = ' '.join(i)
            wordList_1.append(t)
        self.combo_doctor_schedule.addItems(wordList_1)
        self.combo_doctor_schedule.setEditable(True)
        self.combo_doctor_schedule.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.combo_doctor_schedule.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.combo_doctor_schedule.setGeometry(QtCore.QRect(170, 120, 231, 31))
        self.combo_doctor_schedule.setObjectName("combo_doctor_schedule")
        self.combo_patient_schedule = QtWidgets.QComboBox(self.tab_2)

        wordlist = cm.select_table_name('patient').tolist()
        wordList_1 = []
        for i in wordlist:
            t = ' '.join(i)
            wordList_1.append(t)
        self.combo_patient_schedule.addItems(wordList_1)
        self.combo_patient_schedule.setEditable(True)
        self.combo_patient_schedule.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.combo_patient_schedule.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        self.combo_patient_schedule.setGeometry(QtCore.QRect(170, 170, 231, 31))
        self.combo_patient_schedule.setObjectName("combo_patient_schedule")
        self.btn_add_schedule = QtWidgets.QPushButton(self.tab_2)
        self.btn_add_schedule.setGeometry(QtCore.QRect(420, 230, 91, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.btn_add_schedule.setFont(font)
        self.btn_add_schedule.setObjectName("btn_add_schedule")
        self.btn_add_schedule.clicked.connect(self.add_schedule)
        self.date_add_schedule = QtWidgets.QDateTimeEdit(self.tab_2)
        self.date_add_schedule.setDate(datetime.datetime.now())
        self.date_add_schedule.setGeometry(QtCore.QRect(170, 70, 191, 31))
        self.date_add_schedule.setObjectName("date_add_schedule")
        self.tabWidget.addTab(self.tab_2, "")
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 986, 25))
        self.menubar.setObjectName("menubar")
        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.retranslateUi()
        self.tabWidget.setCurrentIndex(0)
        self.init_data(data)
        self.database_config_obj = cm.load_config(cm.DATABASE_CONFIG_PATH)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("schedule_manage", "Schedule Manage"))
        self.label_14.setText(_translate("schedule_manage", "Schedules"))
        item = self.table_schedule.horizontalHeaderItem(0)
        item.setText(_translate("schedule_manage", "ID"))
        item = self.table_schedule.horizontalHeaderItem(1)
        item.setText(_translate("schedule_manage", "Name"))
        item = self.table_schedule.horizontalHeaderItem(2)
        item.setText(_translate("schedule_manage", "Date"))
        item = self.table_schedule.horizontalHeaderItem(3)
        item.setText(_translate("schedule_manage", "Doctor"))
        item = self.table_schedule.horizontalHeaderItem(4)
        item.setText(_translate("schedule_manage", "Patient"))
        item = self.table_schedule.horizontalHeaderItem(5)
        item.setText(_translate("schedule_manage", "Result"))
        item = self.table_schedule.horizontalHeaderItem(6)
        item.setText(_translate("schedule_manage", "Edit"))
        item = self.table_schedule.horizontalHeaderItem(7)
        item.setText(_translate("schedule_manage", "Delete"))
        self.btn_import_schedule.setText(_translate("schedule_manage", "Import"))
        self.btn_export_schedule.setText(_translate("schedule_manage", "Export"))
        self.groupBox.setTitle(_translate("schedule_manage", "Search schedule"))
        self.label_6.setText(_translate("schedule_manage", "Name:"))
        self.label_7.setText(_translate("schedule_manage", "Date From:"))
        self.label_8.setText(_translate("schedule_manage", "Date To:"))
        self.btn_search_schdule.setText(_translate("schedule_manage", "Search Schedule"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("schedule_manage", "Search"))
        self.label_2.setText(_translate("schedule_manage", "Name:"))
        self.label_3.setText(_translate("schedule_manage", "Date:"))
        self.label_4.setText(_translate("schedule_manage", "Doctor ID:"))
        self.label_5.setText(_translate("schedule_manage", "Patient ID:"))
        self.btn_add_schedule.setText(_translate("schedule_manage", "Add"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("schedule_manage", "Add new"))

    def init_data(self, data):
        for i in range(0, len(data)):
            self.table_schedule.insertRow(i)
            self.table_schedule.setItem(i, 0, QtWidgets.QTableWidgetItem(str(data[i][0])))
            self.table_schedule.setItem(i, 1, QtWidgets.QTableWidgetItem(str(data[i][1])))
            self.table_schedule.setItem(i, 2, QtWidgets.QTableWidgetItem(str(data[i][2])))
            self.table_schedule.setItem(i, 3, QtWidgets.QTableWidgetItem(str(data[i][3])))
            self.table_schedule.setItem(i, 4, QtWidgets.QTableWidgetItem(str(data[i][4])))
            self.table_schedule.setItem(i, 5, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.table_schedule.setItem(i, 6, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.table_schedule.setItem(i, 7, QtWidgets.QTableWidgetItem(str(data[i][5])))
            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.edit_btn_click)
            self.detail_btn.setText("Edit")
            self.table_schedule.setCellWidget(i, 6, self.detail_btn)
            self.detail_btn = QtWidgets.QPushButton(self)
            self.detail_btn.clicked.connect(self.delete_btn_click)
            self.detail_btn.setText("Delete")
            self.table_schedule.setCellWidget(i, 7, self.detail_btn)

    def edit_btn_click(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_schedule()
        self.ui.setupUi()
        self.ui.show()
        button = QtWidgets.qApp.focusWidget()
        index = self.table_schedule.indexAt(button.pos())
        table_model = self.table_schedule.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        date_index = table_model.index(index.row(), 2)
        date = table_model.data(date_index)
        doctor_index = table_model.index(index.row(), 3)
        doctor = table_model.data(doctor_index)
        patient_index = table_model.index(index.row(), 4)
        patient = table_model.data(patient_index)
        result_index = table_model.index(index.row(), 5)
        result = table_model.data(result_index)
        self.ui.txt_id_edit_schedule.setText(id)
        self.ui.txt_name_edit_schedule.setText(name)
        date_time_obj = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
        self.ui.date_time_edit_schedule.setDate(date_time_obj)
        self.ui.combo_doctor_edit_schedule.setCurrentText(doctor)
        self.ui.combo_patient_edit_schedule.setCurrentText(patient)
        self.ui.txt_result_edit_schedule.setText(result)

        self.ui.btn_save_edit_schedule.clicked.connect(self.edit_schedule)
        self.ui.btn_cancel_edit_schedule.clicked.connect(self.cancel_edit_schedule)

    def cancel_edit_schedule(self):
        self.ui = QMainWindow()
        self.ui = Ui_edit_schedule()
        self.ui.setupUi()
        self.ui.hide()

    def edit_schedule(self):
        id = self.ui.txt_id_edit_schedule.text()
        name_new = self.ui.txt_name_edit_schedule.text()
        date_new = self.ui.date_time_edit_schedule.dateTime().toPyDateTime()
        doctor_id_new = int(cm.convert_name_to_id(self.ui.combo_doctor_edit_schedule.currentText(), 'doctor'))
        patient_id_new = int(cm.convert_name_to_id(self.ui.combo_patient_edit_schedule.currentText(), 'patient'))
        result_new = self.ui.txt_result_edit_schedule.text()
        if cm.edit_data_schedule(id, name_new, date_new, doctor_id_new, patient_id_new, result_new):
            QMessageBox.information(self, 'Message', f'Successfully edit the schedule: {name_new} ', QMessageBox.Close)
            self.table_schedule.clearContents()
            self.table_schedule.setRowCount(0)
            data_new = cm.select_schedule()
            self.init_data(data_new)
            self.ui = QMainWindow()
            self.ui = Ui_edit_schedule()
            self.ui.setupUi()
            self.ui.hide()
        else:
            QMessageBox.information(self, 'Message', f'Failed edit the schedule: {name_new} ', QMessageBox.Close)
            self.table_schedule.clearContents()
            self.table_schedule.setRowCount(0)
            data_new = cm.select_schedule()
            self.init_data(data_new)

            self.ui = QMainWindow()
            self.ui = Ui_edit_schedule()
            self.ui.setupUi()
            self.ui.hide()

    def delete_btn_click(self):
        button = QtWidgets.qApp.focusWidget()
        index = self.table_schedule.indexAt(button.pos())
        table_model = self.table_schedule.model()
        id_index = table_model.index(index.row(), 0)
        id = table_model.data(id_index)
        name_index = table_model.index(index.row(), 1)
        name = table_model.data(name_index)
        ques = QMessageBox.question(self, 'System', f'Are you sure you want to delete doctor: {name}?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ques == QMessageBox.Yes:
            if cm.delete_data(id, 'schedule'):
                QMessageBox.about(self, "System", f"Successfully delete the schedule: {name}")
                self.table_schedule.clearContents()
                self.table_schedule.setRowCount(0)
                data_new = cm.select_schedule()
                self.init_data(data_new)
            else:
                QMessageBox.about(self, "System", "Delete failed, try again")

    def add_schedule(self):
        name = self.txt_name_schedule.text()
        date = self.date_add_schedule.dateTime().toPyDateTime()
        doctor_id = int(cm.convert_name_to_id(self.combo_doctor_schedule.currentText(), 'doctor'))
        patient_id = int(cm.convert_name_to_id(self.combo_patient_schedule.currentText(), 'patient'))
        if cm.check_empty(name):
            if cm.insert_schedule(name, date, doctor_id, patient_id):
                QMessageBox.information(self, 'System', 'Add successfully', QMessageBox.Close)
                self.table_schedule.clearContents()
                self.table_schedule.setRowCount(0)
                data_new = cm.select_schedule()
                self.init_data(data_new)
            else:
                QMessageBox.information(self, 'Message', 'Add fail', QMessageBox.Close)
                self.table_schedule.clearContents()
                self.table_schedule.setRowCount(0)
                data_new = cm.select_schedule()
                self.init_data(data_new)
        else:
            QMessageBox.information(self, 'Message', 'Name is not null', QMessageBox.Close)


    def search_schedule(self):
        name = self.txt_search_schedule.text()
        date_from = self.data_from_search_schedule.dateTime().toPyDateTime()
        date_to = self.date_to_search_schedule.dateTime().toPyDateTime()
        data_schedule = cm.search_schedule(date_from, date_to, name)
        if len(data_schedule) != 0:
            self.table_schedule.clearContents()
            self.table_schedule.setRowCount(0)
            self.init_data(data_schedule)
        else:
            QMessageBox.information(self, 'Message', 'Nothing found!', QMessageBox.Close)

    def import_schedule(self):
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename()

        if 'xlsx' not in file_path:
            QMessageBox.information(self, 'System', 'Please import into excel file!', QMessageBox.Close)
        else:
            wb = load_workbook(file_path)
            ws = wb.worksheets[0]
            list_data = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    list_data.append(cell.value)
            list_data_result = list(cm.divide_chunks(list_data, ws.max_column))
            for i in list_data_result:
                for j in range(0, len(i), ws.max_column):
                    cm.insert_schedule(i[j], i[j + 1], i[j + 2], i[j + 3])
            self.table_schedule.clearContents()
            self.table_schedule.setRowCount(0)
            data_new = cm.select_schedule()
            self.init_data(data_new)
            QMessageBox.information(self, 'System', 'Import successfully', QMessageBox.Close)

    def export_schedule(self):
        columnHeaders = []
        for j in range(self.table_schedule.model().columnCount() - 2):
            columnHeaders.append(self.table_schedule.horizontalHeaderItem(j).text())

        df = pd.DataFrame(columns=columnHeaders)
        for row in range(self.table_schedule.rowCount()):
            for col in range(self.table_schedule.columnCount() - 2):
                df.at[row, columnHeaders[col]] = self.table_schedule.item(row, col).text()
        wb = Workbook()
        wb = load_workbook('Template/template_schedule_export.xlsx')
        ws1 = wb.worksheets[0]
        offset_row = 1
        offset_col = 0
        row = 1
        for row_data in dataframe_to_rows(df, index=False, header=False):
            col = 1
            for cell_data in row_data:
                ws1.cell(row + offset_row, col + offset_col, cell_data)
                col += 1
            row += 1
        if os.path.exists('Output') == False:
            folder_path = 'Output'
            os.mkdir(folder_path)
        (d, m, y, h, mi, s) = cm.split_date_time()
        wb.save(f'C:\\Users\ADMIN\PycharmProjects\DucBNN_Mock_Project\Output\Schedule_{d}_{m}_{y}_{h}_{mi}_{s}.xlsx')


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = Ui_schedule_manage()
    ui.setupUi()
    ui.show()
    sys.exit(app.exec_())
